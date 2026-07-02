"""
Transform a raw "OTA discount grid" xlsx (the side-by-side INTERNATIONAL/DOMESTIC
dump) into a clean, print-ready report.

Detects layout (does not assume fixed cells), validates assumptions, fails loud,
preserves the raw input (writes a NEW *_report.xlsx). See the task spec for rules.

Usage:
  python tools/format_discount_report.py <raw.xlsx> [--out <report.xlsx>]
"""
from __future__ import annotations

import argparse
import re
import sys
from datetime import date
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import PageSetupProperties

# --- styling constants -------------------------------------------------------
SUPERSCRIPTS = ["ᵃ", "ᵇ", "ᶜ", "ᵈ", "ᵉ", "ᶠ",
                "ᵍ", "ʰ", "ⁱ", "ʲ", "ᵏ", "ˡ",
                "ᵐ", "ⁿ", "ᵒ", "ᵖ"]  # ᵃ ᵇ ᶜ ᵈ ᵉ ᶠ ...
NUMFMT = '0.00"%";[Red](0.00"%")'
TITLE_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FILL = PatternFill("solid", fgColor="2E5496")
LABEL_FILL = PatternFill("solid", fgColor="D9E1F2")
DIVIDER_FILL = PatternFill("solid", fgColor="BDD7EE")
BESTRATE_FILL = PatternFill("solid", fgColor="1F3864")
BESTOTA_FILL = PatternFill("solid", fgColor="BDD7EE")
WIN_FILL = PatternFill("solid", fgColor="C6EFCE")
WIN_FONT = Font(color="006100", bold=True)
WHITE_BOLD = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
MONTHS = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def die(msg: str) -> None:
    raise SystemExit(f"FAIL: {msg}")


def warn(msg: str) -> None:
    print(f"  WARN: {msg}")


# --- numeric helpers ---------------------------------------------------------
NUM_RE = re.compile(r"-?\d+\.?\d*")
PAIR_RE = re.compile(r"(-?\d+\.?\d*)\s*(?:\(([^)]+)\))?")
PAREN_RE = re.compile(r"\(([^)]+)\)")


def as_float(value: Any) -> Optional[float]:
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def classify(value: Any) -> str:
    """'blank' | 'numeric' | 'card' | 'other'."""
    if value is None or str(value).strip() == "":
        return "blank"
    if as_float(value) is not None:
        return "numeric"
    s = str(value)
    if PAREN_RE.search(s) and NUM_RE.search(s):
        nums = NUM_RE.findall(s)
        return "card" if 1 <= len(nums) <= 2 else "other"
    return "other"


def cell_best_number(value: Any) -> Optional[float]:
    """Numeric value, or MAX of numbers in a card cell, or None for blank/other."""
    kind = classify(value)
    if kind == "numeric":
        return as_float(value)
    if kind == "card":
        return max(float(n) for n in NUM_RE.findall(str(value)))
    return None


# --- layout detection --------------------------------------------------------
def detect_tables(ws) -> tuple[list[dict], int]:
    ota_cells = [(c.row, c.column) for row in ws.iter_rows()
                 for c in row if str(c.value).strip() == "OTA"]
    if len(ota_cells) != 2:
        die(f"expected exactly 2 'OTA' headers, found {len(ota_cells)}")
    rows = {r for r, _ in ota_cells}
    if len(rows) != 1:
        die(f"the two 'OTA' headers are on different rows {sorted(rows)}")
    header_row = ota_cells[0][0]
    ota_cols = sorted(c for _, c in ota_cells)

    tables = []
    for ota_col in ota_cols:
        # airline columns: contiguous non-empty header cells to the right
        airlines = []
        col = ota_col + 1
        while col <= ws.max_column:
            v = ws.cell(header_row, col).value
            if v is None or str(v).strip() == "":
                break
            airlines.append((col, str(v).strip()))
            col += 1
        if not airlines:
            die(f"no airline columns found right of OTA at column {ota_col}")

        # OTA rows downward until two consecutive blank label rows / end
        labels, blanks, r, consec_blank = [], [], header_row + 1, 0
        while r <= ws.max_row + 1:
            v = ws.cell(r, ota_col).value
            if v is None or str(v).strip() == "":
                consec_blank += 1
                if consec_blank >= 2:
                    break
                blanks.append(r)
            else:
                consec_blank = 0
                labels.append((r, str(v).strip()))
            r += 1
        # the single blank inside the block (between first and last label) = divider
        last_label = labels[-1][0]
        inner_blanks = [b for b in blanks if header_row < b < last_label]
        if len(inner_blanks) != 1:
            die(f"expected exactly 1 divider blank row in OTA block at col {ota_col}, "
                f"found {len(inner_blanks)} ({inner_blanks})")
        tables.append({
            "ota_col": ota_col,
            "header_row": header_row,
            "airlines": airlines,                      # [(col, code)]
            "ota_rows": labels,                        # [(row, label)] (excludes divider)
            "divider_row": inner_blanks[0],
            "title_cell": (header_row - 1, ota_col),
        })

    # spacer column = the empty column just left of the right table's OTA col
    spacer = tables[1]["ota_col"] - 1
    if any(str(ws.cell(header_row, spacer).value or "").strip() for _ in [0]):
        die(f"expected blank spacer column at {spacer}, but it has a header")

    # validation: identical OTA name lists in identical order
    left = [lbl for _, lbl in tables[0]["ota_rows"]]
    right = [lbl for _, lbl in tables[1]["ota_rows"]]
    if left != right:
        die(f"OTA lists differ between tables:\n  left={left}\n  right={right}")
    return tables, spacer


# --- title parsing -----------------------------------------------------------
TITLE_RE = re.compile(r"(\d{2})/(\d{2})/(\d{4}).*?\(([^)]+)\).*?(\d{3,4})\s*hrs", re.I)


def render_title(raw: Any) -> str:
    s = str(raw or "").strip()
    m = TITLE_RE.search(s)
    if not m:
        cleaned = re.sub(r"[()—–-]", " ", s).strip()
        warn(f"could not parse title {s!r}; using fallback {cleaned!r}")
        return re.sub(r"\s+", " ", cleaned)
    dd, mm, yyyy, label, hhmm = m.groups()
    try:
        d = date(int(yyyy), int(mm), int(dd))
        datepart = f"{d.day} {MONTHS[d.month]} {d.year}"
    except ValueError:
        warn(f"bad date in title {s!r}; using raw digits")
        datepart = f"{dd}/{mm}/{yyyy}"
    return f"{label.strip().title()}   {datepart} · {hhmm} hrs"


# --- abbreviation key --------------------------------------------------------
def build_abbreviations(labels: list[str]) -> dict[str, str]:
    """Short (<=8 char) unique tags. The -B/-C persona tag is added ONLY when the
    same base appears as both B2B and B2C (a real collision); otherwise omitted."""
    filler = {"OTA", "AIR", "AIRLINE", "AIRLINES", "AIRWAYS", "LIMITED", "LTD"}

    def split(name: str) -> tuple[str, Optional[str]]:
        m = re.search(r"-?\s*B2([BC])\b", name, re.I)
        persona = m.group(1).upper() if m else None
        base = re.sub(r"[-\s]*B2[BC]\b", "", name, flags=re.I).strip()
        tokens = [tok for tok in re.split(r"[\s\-]+", base) if tok and tok.upper() not in filler]
        base = " ".join(tokens) if tokens else base
        return base, persona

    personas: dict[str, set[str]] = {}
    for name in labels:
        base, persona = split(name)
        if persona:
            personas.setdefault(base.upper(), set()).add(persona)
    collide = {b for b, ps in personas.items() if len(ps) > 1}

    out: dict[str, str] = {}
    used: set[str] = set()
    for name in labels:
        base, persona = split(name)
        core = re.sub(r"[^A-Za-z0-9]", "", base) or "OTA"
        tag = f"-{persona}" if (persona and base.upper() in collide) else ""
        cand = core[: 8 - len(tag)] + tag
        final, i = cand, 1
        while final in used:
            final = f"{cand[:7]}{i}"
            i += 1
        used.add(final)
        out[name] = final
    return out


# --- main transform ----------------------------------------------------------
def transform(in_path: Path, out_path: Path) -> None:
    wb = openpyxl.load_workbook(in_path)
    ws = wb.active
    tables, spacer = detect_tables(ws)
    print("Detected:")
    for t in tables:
        print(f"  table @col{t['ota_col']}: {len(t['airlines'])} airlines, "
              f"{len(t['ota_rows'])} OTAs, divider row {t['divider_row']}")
    print(f"  spacer column: {spacer}")

    header_row = tables[0]["header_row"]
    ota_rows = tables[0]["ota_rows"]
    divider_row = tables[0]["divider_row"]
    first_data_row, last_data_row = ota_rows[0][0], ota_rows[-1][0]
    last_report_col = max(c for c, _ in tables[1]["airlines"])

    # --- card markers (scan whole grid; track which cards appear per table) --
    card_marker: dict[str, str] = {}
    card_order: list[str] = []
    table_cards: list[set[str]] = [set() for _ in tables]
    for ti, t in enumerate(tables):
        for col, _ in t["airlines"]:
            for r, _ in t["ota_rows"]:
                v = ws.cell(r, col).value
                if classify(v) == "card":
                    for c in PAREN_RE.findall(str(v)):
                        c = c.strip()
                        if c not in card_marker:
                            if len(card_order) >= len(SUPERSCRIPTS):
                                warn(f"more cards than markers; '{c}' unmarked")
                                continue
                            card_marker[c] = SUPERSCRIPTS[len(card_order)]
                            card_order.append(c)
                        if c in card_marker:
                            table_cards[ti].add(c)

    # --- titles --------------------------------------------------------------
    for t in tables:
        tr, tc = t["title_cell"]
        last_col = max(c for c, _ in t["airlines"])
        ws.cell(tr, tc).value = render_title(ws.cell(tr, tc).value)
        ws.merge_cells(start_row=tr, start_column=tc, end_row=tr, end_column=last_col)
        a = ws.cell(tr, tc)
        a.fill, a.font, a.alignment = TITLE_FILL, Font(color="FFFFFF", bold=True, size=13), CENTER
        ws.row_dimensions[tr].height = 20

    # --- headers + OTA label column ------------------------------------------
    for t in tables:
        h = ws.cell(header_row, t["ota_col"])
        h.fill, h.font, h.alignment, h.border = HEADER_FILL, WHITE_BOLD, CENTER, BORDER
        for col, _ in t["airlines"]:
            c = ws.cell(header_row, col)
            c.fill, c.font, c.alignment, c.border = HEADER_FILL, WHITE_BOLD, CENTER, BORDER
        for r, _ in t["ota_rows"]:
            c = ws.cell(r, t["ota_col"])
            c.fill, c.font, c.alignment, c.border = LABEL_FILL, Font(bold=True), LEFT, BORDER

    # --- data grid -----------------------------------------------------------
    for t in tables:
        for col, _ in t["airlines"]:
            for r, _ in t["ota_rows"]:
                cell = ws.cell(r, col)
                kind = classify(cell.value)
                cell.border = BORDER
                cell.alignment = CENTER
                if kind == "blank":
                    continue
                if kind == "numeric":
                    cell.value = as_float(cell.value)
                    cell.number_format = NUMFMT
                elif kind == "card":
                    pairs = [(float(num), (card.strip() if card else None))
                             for num, card in PAIR_RE.findall(str(cell.value)) if num]
                    pairs.sort(key=lambda p: p[0])
                    parts = [f"{num:g}{card_marker.get(card, '') if card else ''}"
                             for num, card in pairs]
                    cell.value = " / ".join(parts)
                else:
                    warn(f"cell {cell.coordinate}={cell.value!r} unrecognized; left unchanged")

    # --- divider row ---------------------------------------------------------
    for t in tables:
        for col in [t["ota_col"]] + [c for c, _ in t["airlines"]]:
            ws.cell(divider_row, col).fill = DIVIDER_FILL
    ws.row_dimensions[divider_row].height = 6

    # --- helper grid + abbreviation key (grouped, collapsed, right of report) -
    key_col = last_report_col + 2
    ws.cell(header_row, key_col).value = "Best-OTA key"
    ws.cell(header_row, key_col).font = Font(bold=True, size=9)
    abbr_map = build_abbreviations([lbl for _, lbl in ota_rows])
    abbr_by_row: dict[int, str] = {}
    for r, label in ota_rows:
        ab = abbr_map[label]
        abbr_by_row[r] = ab
        c = ws.cell(r, key_col)
        c.value, c.font = ab, Font(italic=True, size=9)

    # one helper column per airline column (both tables); map report_col -> helper_col.
    # Numeric cells are floats now; card cells are text "lo / hi" -> parse the max number.
    helper_col_of: dict[int, int] = {}
    hcol = key_col + 1
    for t in tables:
        for col, _ in t["airlines"]:
            helper_col_of[col] = hcol
            for r, _ in t["ota_rows"]:
                val = ws.cell(r, col).value
                num = as_float(val)
                if num is None and val is not None and str(val).strip():
                    nums = NUM_RE.findall(str(val))
                    num = max(float(x) for x in nums) if nums else None
                if num is not None:
                    ws.cell(r, hcol).value = num
            hcol += 1
    last_helper_col = hcol - 1

    # collapse key + helper columns (outline level 1, hidden)
    for col in range(key_col, last_helper_col + 1):
        cd = ws.column_dimensions[get_column_letter(col)]
        cd.outline_level = 1
        cd.hidden = True
    ws.sheet_properties.outlinePr.summaryRight = True

    # --- Best rate + Best OTA rows (computed values + formula in a comment) ---
    best_rate_row = last_data_row + 1
    best_ota_row = best_rate_row + 1
    rng_rows = (first_data_row, last_data_row)
    for t in tables:
        rate_label = ws.cell(best_rate_row, t["ota_col"])
        ota_label = ws.cell(best_ota_row, t["ota_col"])
        rate_label.value, ota_label.value = "Best rate", "Best OTA"
        rate_label.fill, rate_label.font = BESTRATE_FILL, Font(bold=True, color="FFFFFF")
        ota_label.fill, ota_label.font = BESTOTA_FILL, Font(bold=True, color="1F3864")
        # one formula note per row (auditability) instead of cluttering every cell
        rate_label.comment = Comment(
            'Per airline = IF(COUNT(helper_col)=0,"–",MAX(helper_col)) over the hidden '
            'numeric helper grid (captures card-condition highs).', "report")
        ota_label.comment = Comment(
            'Per airline = IF(COUNT(helper_col)=0,"–",INDEX(key,MATCH(MAX(helper_col),'
            'helper_col,0))) using the hidden abbreviation key.', "report")
        for col, _ in t["airlines"]:
            hcol = helper_col_of[col]
            # python-computed values (headless-safe; matches the helper-grid formula)
            vals = [ws.cell(r, hcol).value for r, _ in t["ota_rows"]]
            present = [(v, r) for (r, _), v in zip(t["ota_rows"], vals) if isinstance(v, (int, float))]
            br = ws.cell(best_rate_row, col)
            bo = ws.cell(best_ota_row, col)
            if not present:
                br.value, bo.value = "–", "–"
            else:
                best = max(present, key=lambda p: p[0])
                br.value = best[0]
                br.number_format = NUMFMT
                bo.value = abbr_by_row[best[1]]
            for c in (br, bo):
                c.alignment, c.border = CENTER, BORDER
            br.fill, br.font = BESTRATE_FILL, Font(bold=True, color="FFFFFF")
            bo.fill, bo.font = BESTOTA_FILL, Font(bold=True, color="1F3864", size=8)
    ws.row_dimensions[best_ota_row].height = 14

    # --- conditional formatting: highlight numeric per-column max (>0) --------
    for t in tables:
        for col, _ in t["airlines"]:
            cl = get_column_letter(col)
            rng = f"{cl}{first_data_row}:{cl}{last_data_row}"
            rule = FormulaRule(
                formula=[f"AND({cl}{first_data_row}=MAX({cl}${first_data_row}:{cl}${last_data_row}),"
                         f"{cl}{first_data_row}>0)"],
                fill=WIN_FILL, font=WIN_FONT)
            ws.conditional_formatting.add(rng, rule)

    # --- footnotes (one per card, under each table) --------------------------
    foot_start = best_ota_row + 2
    last_foot_row = foot_start - 1
    for ti, t in enumerate(tables):
        last_col = max(c for c, _ in t["airlines"])
        fr = foot_start
        # only cards that actually appear in THIS table, in global marker order
        for card in [c for c in card_order if c in table_cards[ti]]:
            ws.cell(fr, t["ota_col"]).value = (
                f"{card_marker[card]}  Rates shown as standard / card rate ({card}). "
                f"Values are discount %.")
            ws.merge_cells(start_row=fr, start_column=t["ota_col"], end_row=fr, end_column=last_col)
            fc = ws.cell(fr, t["ota_col"])
            fc.font = Font(italic=True, size=9, color="595959")
            fc.alignment = LEFT
            fr += 1
        last_foot_row = max(last_foot_row, fr - 1)

    # --- widths / heights ----------------------------------------------------
    for t in tables:
        ws.column_dimensions[get_column_letter(t["ota_col"])].width = 16
    intl_cols = [c for c, _ in tables[0]["airlines"]]
    dom_cols = [c for c, _ in tables[1]["airlines"]]
    for c in intl_cols:
        ws.column_dimensions[get_column_letter(c)].width = 5.5
    for c in dom_cols:
        ws.column_dimensions[get_column_letter(c)].width = 6.5
    ws.column_dimensions[get_column_letter(spacer)].width = 3.5
    for r, _ in ota_rows:
        ws.row_dimensions[r].height = 15

    # --- print setup (fit 1x1 landscape, exclude grouped cols) ---------------
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = True
    ws.print_area = f"A1:{get_column_letter(last_report_col)}{last_foot_row}"
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.4, bottom=0.25)
    wb.calculation.fullCalcOnLoad = True

    # --- self-check ----------------------------------------------------------
    print("\nSelf-check:")
    problems = 0
    for t in tables:
        for col, code in t["airlines"]:
            hcol = helper_col_of[col]
            vals = [ws.cell(r, hcol).value for r, _ in t["ota_rows"]]
            nums = [v for v in vals if isinstance(v, (int, float))]
            br = ws.cell(best_rate_row, col).value
            if not nums:
                if br != "–":
                    warn(f"col {code}: all-blank should be '–', got {br!r}"); problems += 1
            elif br != max(nums):
                warn(f"col {code}: best-rate {br} != helper max {max(nums)}"); problems += 1
    plain = "abcdefghijklmnop"
    print(f"  cards: {len(card_order)} -> "
          + ", ".join(f"[{plain[i]}]={c}" for i, c in enumerate(card_order)))
    print(f"  best-rate/best-OTA rows: {best_rate_row}/{best_ota_row}; footnotes end {last_foot_row}")
    print(f"  helper+key columns {get_column_letter(key_col)}..{get_column_letter(last_helper_col)} (collapsed)")
    print(f"  {'OK' if problems == 0 else str(problems) + ' PROBLEM(S)'}")

    wb.save(out_path)
    print(f"\nSaved report -> {out_path}")


def main() -> int:
    ap = argparse.ArgumentParser(description="Format a raw OTA discount grid xlsx into a print-ready report.")
    ap.add_argument("input", help="raw .xlsx file")
    ap.add_argument("--out", default=None, help="output path (default <input>_report.xlsx)")
    args = ap.parse_args()
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:  # noqa: BLE001
        pass
    in_path = Path(args.input)
    if not in_path.exists():
        die(f"input not found: {in_path}")
    out_path = Path(args.out) if args.out else in_path.with_name(in_path.stem + "_report.xlsx")
    transform(in_path, out_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
