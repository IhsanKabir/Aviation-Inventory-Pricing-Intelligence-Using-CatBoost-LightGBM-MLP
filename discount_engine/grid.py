"""
OTA Discount GRID engine (the "25 June" format).

Builds the channel x airline discount grid, split DOMESTIC / INTERNATIONAL,
from the per-channel discount extractors:

  * FirstTrip B2C   -> live  (modules.firsttrip.fetch_b2c_discounts)        single %
  * FirstTrip B2B   -> HAR   (modules.firsttrip.parse_b2b_commissions)      agent commission
  * GoZayaan B2C    -> HAR   (modules.gozayaan_har.parse_discounts)         common(payment), special(card)
  * ShareTrip B2C   -> HAR   (modules.sharetrip_har)                        common(bKash), special(card)
  * AmyBD / AKIJ / BDFare -> HAR (modules.amyweb / akijair_har / bdfare_har) agent commission

This is the LIBRARY: import `discount_engine` and call `build_report(...)` /
`run_report(...)`. The CLI lives in `tools/ota_discount_grid.py` (a thin wrapper).

True base (DEFAULT ON): BDFare and AKIJ report base differently (BDFare estimates
base = gross*ratio; AKIJ reclassifies part of base as tax), which skews their % off
base. By default the grid recomputes their DOMESTIC cells on the canonical base
learned from the exact-base channels (FT B2B/B2C) so the report shows the ACTUAL %.
INTL is unaffected (intl tax varies). See tools/base_fare_audit.py for the audit.
"""

from __future__ import annotations

import csv
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from modules import firsttrip, gozayaan_har, amyweb, sharetrip_har, akijair_har, bdfare_har

from .highlight import (
    BEST_SHORT,
    compute_highlights,
    leading_number as _leading_number,
    prev_lookup_from_report,
)

# Bangladesh domestic airports — a route is DOMESTIC iff both ends are here.
DOMESTIC_AIRPORTS = {"DAC", "CGP", "CXB", "ZYL", "SPD", "BZL", "RJH", "JSR", "SAH", "TKR", "IRD", "KMI"}

# Row order mirrors the manual report (B2B block then B2C block).
ROW_ORDER = [
    ("USBA OTA B2B", "b2b"),
    ("SHARETRIP-B2B", "b2b"),
    ("BDFare", "b2b"),
    ("TLN", "b2b"),
    ("AKIJ AIR-B2B", "b2b"),
    ("__sep__", "sep"),
    ("Firsttrip-B2C", "b2c"),
    ("ShareTrip-B2C", "b2c"),
    ("Go Zayaan", "b2c"),
    ("Amy", "b2c"),
]

DOM_COLUMNS = ["BS", "2A", "BG", "VQ"]
INTL_COLUMNS = ["BS", "BG", "EK", "SV", "QR", "SQ", "MH", "CA", "AI", "EY", "GF", "KU",
                "OD", "TG", "TK", "UL", "WY", "CX", "CZ", "H9", "MU", "ET", "MS", "6E",
                "G9", "FZ", "J9", "8D"]


def _route_type(origin: str, destination: str) -> str:
    return "DOM" if origin in DOMESTIC_AIRPORTS and destination in DOMESTIC_AIRPORTS else "INTL"


def _fmt(value: float) -> str:
    """Trim trailing .0 (16.0 -> '16', 8.7 -> '8.7')."""
    return f"{value:g}"


# --- per-channel collectors: return cell dicts keyed by (route_type, airline) -----------

def _fetch_firsttrip_b2c(routes: list[tuple[str, str, Optional[str]]],
                         default_date: Optional[str],
                         ) -> dict[tuple[str, str, str], list[dict[str, Any]]]:
    """Live-fetch FirstTrip B2C rows ONCE per route: {(origin, dest, date): rows}.
    Shared by the true-base oracle and the B2C collector — no double fetch."""
    rows_by_route: dict[tuple[str, str, str], list[dict[str, Any]]] = {}
    for origin, destination, route_date in (routes or []):
        date = route_date or default_date
        if not date:
            print(f"  ! FirstTrip B2C {origin}-{destination}: no date "
                  f"(use --date or DAC-CGP@YYYY-MM-DD); skipped")
            continue
        try:
            rows_by_route[(origin, destination, date)] = firsttrip.fetch_b2c_discounts(
                origin, destination, date)
        except Exception as exc:  # noqa: BLE001
            print(f"  ! FirstTrip B2C {origin}-{destination}: {exc}")
    return rows_by_route


def _collect_firsttrip_b2c_rows(
        rows_by_route: dict[tuple[str, str, str], list[dict[str, Any]]],
        ) -> dict[tuple[str, str], str]:
    cells: dict[tuple[str, str], str] = {}
    for (origin, destination, _date), rows in rows_by_route.items():
        rt = _route_type(origin, destination)
        summary = firsttrip.summarize_b2c_discounts(rows)
        for airline, cell in summary.items():
            key = (rt, airline)
            # keep the highest rate seen for this airline/route_type
            text = _fmt(cell["rate"])
            if key not in cells or cell["rate"] > _existing_rate(cells[key]):
                cells[key] = text
    return cells


def collect_firsttrip_b2c(routes: list[tuple[str, str, Optional[str]]],
                          default_date: Optional[str]) -> dict[tuple[str, str], str]:
    """Fetch + summarize in one call (kept for direct use; build_report shares rows)."""
    return _collect_firsttrip_b2c_rows(_fetch_firsttrip_b2c(routes, default_date))


def _existing_rate(text: str) -> float:
    try:
        return float(text.split(",")[0].split()[0])
    except (ValueError, IndexError):
        return -1.0


def collect_gozayaan(har_path: str) -> dict[tuple[str, str], str]:
    cells: dict[tuple[str, str], str] = {}
    rows = gozayaan_har.parse_discounts(har_path)
    summary = gozayaan_har.summarize_discounts(rows)  # keyed (airline, flight_type)
    for (airline, flight_type), cell in summary.items():
        rt = "DOM" if flight_type == "DOM" else "INTL"
        common = cell.get("common_pct")
        if common is None:
            continue
        text = _fmt(common)
        special = cell.get("special")
        if special:
            text += f", {_fmt(special['pct'])} ({special['eligibility']})"
        cells[(rt, airline)] = text
    return cells


def _collect_firsttrip_b2b_rows(rows: list[dict[str, Any]]) -> dict[tuple[str, str], str]:
    # representative per (airline, route_type) = highest commission
    best: dict[tuple[str, str], float] = {}
    for r in rows:
        rt = _route_type(r["origin"], r["destination"])
        key = (rt, r["airline"])
        if key not in best or r["commission_pct"] > best[key]:
            best[key] = r["commission_pct"]
    return {key: _fmt(pct) for key, pct in best.items()}


def collect_firsttrip_b2b(har_path: str) -> dict[tuple[str, str], str]:
    """Parse + summarize in one call (kept for direct use; build_report shares rows)."""
    return _collect_firsttrip_b2b_rows(firsttrip.parse_b2b_commissions(har_path))


def _sharetrip_cell_text(c: dict[str, Any]) -> str:
    """`common(wallet[, fee]), special (label[, capped][, fee])[, best card ...]`.

    Percentages are gross of the gateway convenience fee (comparable with other
    channels) but each shown option is annotated with its cheapest eligible
    gateway fee, and the judge already RANKS options net of that fee. The
    leading number stays the common rate so highlight ranking is unchanged.
    """
    def _fee(pct: Any) -> str:
        return f", {_fmt(float(pct))}% fee" if pct else ""

    wallet = ""
    if c.get("common_code"):
        name = "Nagad" if "nagad" in str(c["common_code"]).lower() else "Bkash"
        wallet = f"({name}{_fee(c.get('common_fee_pct'))})"
    text = _fmt(c["common_pct"]) + wallet
    if c.get("special_pct") is not None:
        cap = ", capped" if c.get("special_capped") else ""
        text += f", {_fmt(c['special_pct'])} ({c['special_label']}{cap}{_fee(c.get('special_fee_pct'))})"
    if c.get("card_pct") is not None:
        cap = ", capped" if c.get("card_capped") else ""
        text += f", {_fmt(c['card_pct'])} ({c['card_label']}{cap}{_fee(c.get('card_fee_pct'))})"
    return text


def collect_sharetrip_b2c(har_paths: str | list[str]) -> dict[tuple[str, str], str]:
    """Build ShareTrip cells from one or more HARs, judging ALL coupons.

    Booking-details captures are the reliable source (browser HAR exports evict
    the multi-MB search response bodies). Every coupon in a details payload is
    judged CAP-AWARE at the observed fare — effective saving = min(pct x base,
    maximumDiscountAmount), plus the automatic discount when the coupon stacks —
    so an "18%" card capped at 6,000 BDT on a 91k itinerary correctly loses to a
    1% uncapped loyalty stack. Cells show the common rate (automatic + wallet
    stack), the best judged special, and the best card special when different.

    Coupon TERMS are market-uniform (verified 2026-07-06: identical coupon
    objects across airlines within DOM / within INTL), so airlines that only
    appear in a search capture are judged with the shared terms at their own
    observed base fare. Only displayPrice.discount is airline-specific.
    """
    paths = [har_paths] if isinstance(har_paths, str) else list(har_paths)

    # 1. Exact per-airline cells: judge every booking-flow details payload.
    #    All rows are summarized together so a second HAR can't silently
    #    overwrite a better cell (best common + best judged special win).
    all_rows: list[dict[str, Any]] = []
    for har_path in paths:
        all_rows += sharetrip_har.parse_details_discounts(har_path)
    details = sharetrip_har.summarize_details(all_rows)

    # Market-level coupon terms (uniform per DOM/INTL) + the gateway fee catalog
    # from any booking capture.
    shared_terms: dict[str, list[dict[str, Any]]] = {}
    for r in all_rows:
        rt = "DOM" if r["flight_type"] == "DOM" else "INTL"
        if r.get("coupon_terms"):
            shared_terms.setdefault(rt, r["coupon_terms"])
    shared_gateways: dict[str, dict[str, Any]] = {}
    for har_path in paths:
        shared_gateways.update(sharetrip_har.parse_payment_gateways(har_path))

    cells: dict[tuple[str, str], str] = {}
    for (airline, flight_type), c in details.items():
        rt = "DOM" if flight_type == "DOM" else "INTL"
        cells[(rt, airline)] = _sharetrip_cell_text(c)

    # 2. Search fill: airlines without a booking capture get the shared market
    #    terms judged at their own observed fare (caps re-evaluated per airline;
    #    fees annotated from the shared catalog — net ranking needs the total
    #    fare, which search rows don't carry).
    for har_path in paths:
        common = sharetrip_har.summarize_discounts(sharetrip_har.parse_discounts(har_path))
        for (airline, flight_type), cell in common.items():
            rt = "DOM" if flight_type == "DOM" else "INTL"
            if (rt, airline) in cells:
                continue    # a booking capture already gave the exact cell
            judged = sharetrip_har.judge_cell(cell["discount_pct"],
                                              float(cell.get("base_fare_bdt") or 0),
                                              shared_terms.get(rt) or [],
                                              gateways=shared_gateways or None)
            cells[(rt, airline)] = _sharetrip_cell_text(judged)
    return cells


def collect_akij(har_path: str, field: str, true_base=None) -> dict[tuple[str, str], str]:
    rows = [dict(r) for r in akijair_har.parse_commissions(har_path)]  # copy: we may rewrite %
    if true_base is not None and field == "realized_discount_pct":
        # Unified model: discount = (actual market gross - AKIJ net total) / actual base.
        # true_base.discount() trusts AKIJ's own gross when it shows a real markdown, and
        # marks the net down from the market gross when AKIJ hid it (gross==total, e.g. BG).
        # DROP unresolvable domestic rows so they can't win summarize()'s max().
        kept = []
        for r in rows:
            if _route_type(r["origin"], r["destination"]) != "DOM":
                kept.append(r)
                continue
            pct, _g, _b = true_base.discount(r["airline"], r["gross_fare_bdt"], r["total_fare_bdt"])
            if pct is None:
                continue
            r["realized_discount_pct"] = pct
            kept.append(r)
        rows = kept
    summary = akijair_har.summarize_commissions(rows, field=field)
    return {key: _fmt(cell["value"]) for key, cell in summary.items()}


# Base sources honest enough to show without the '~' estimate marker: a solid
# cross-source match ("market"), BDFare's own captured Fare Summary ("exact"),
# the SAME airline's own tax on this route ("airline_route_tax" — tax is fixed
# per airline+route), and the domestic true-base oracle rewrite ("true_base").
# "route_tax" (another airline's same-route tax) stays an estimate ('~').
_BDFARE_SOLID_BASES = {"market", "exact", "airline_route_tax", "true_base"}


def _market_base_index(row_lists: list[tuple[str, list[dict[str, Any]]]]) -> dict[tuple, tuple]:
    """fare_match_key -> (base_fare, source) from solid B2B sources (Amy, USBA
    FT-B2B). The same flight at the same gross fare is the same published fare,
    so its base is authoritative — BDFare's search list carries no base at all
    and its own Fare Summary is cross-checked against this."""
    index: dict[tuple, tuple] = {}
    for source, rows in row_lists:
        for r in rows:
            base = r.get("base_fare") or r.get("base_fare_bdt")
            gross = r.get("tot_fare") or r.get("gross_total_bdt")
            dep = str(r.get("departure") or "")
            if not base or not gross or len(dep) < 16:
                continue
            try:
                dt = datetime.fromisoformat(dep.replace(" ", "T"))
            except ValueError:
                continue
            key = bdfare_har.fare_match_key(
                r.get("airline", ""), r.get("origin", ""), r.get("destination", ""),
                dt.day, dt.strftime("%b"), dt.strftime("%H:%M"), gross)
            index.setdefault(key, (float(base), source))
    return index


def collect_bdfare(har_path: str, true_base=None,
                   base_index: Optional[dict[tuple, tuple]] = None) -> dict[tuple[str, str], str]:
    rows = [dict(r) for r in bdfare_har.parse_commissions(har_path, base_index=base_index)]
    if true_base is not None:
        # Unified model: agent discount = (actual market gross - agentAmount) / actual base
        # (the agent's discount off the public price, consistent with every other channel).
        # DROP unresolvable domestic rows so they can't win summarize()'s max(); intl offers
        # pass through unchanged (no domestic true base).
        kept = []
        for r in rows:
            if not r["domestic"]:
                kept.append(r)
                continue
            pct, _g, _b = true_base.discount(r["airline"], r["gross_bdt"], r["agent_bdt"])
            if pct is None:
                continue
            r["commission_pct"] = pct
            r["base_source"] = "true_base"   # oracle-normalized = solid, no '~'
            kept.append(r)
        rows = kept
    summary = bdfare_har.summarize_commissions(rows)
    for (rt, airline), cell in sorted(summary.items()):
        if cell.get("n_offers", 1) > 1 and cell.get("pct_max") != cell.get("value"):
            # Provenance for the run log: which fare the cell reflects, and the
            # spread across the other (usually premium) offers in the capture.
            print(f"BDFare {airline} {rt}: {cell['value']}% from the cheapest offer "
                  f"({cell.get('offer_route')} gross {cell.get('offer_gross_bdt'):,}) - "
                  f"{cell['n_offers']} offers ranged "
                  f"{cell.get('pct_min')}-{cell.get('pct_max')}%")
        if cell.get("base_source") not in _BDFARE_SOLID_BASES:
            # An estimated base misstates airlines whose tax share differs from
            # the assumed ratio (AI showed 5.74% when the true figure was 7.49%).
            print(f"BDFare {airline} {rt}: ~{cell['value']}% uses an ESTIMATED base "
                  f"({cell.get('base_source')}) — capture the same flight on Amy/USBA, "
                  f"or open Flight Details -> Fare Summary on {airline}'s cheapest fare, "
                  f"for the exact %")
    # Estimated bases render as '~x.x' — never presented as solid numbers.
    return {key: (("" if cell.get("base_source") in _BDFARE_SOLID_BASES else "~")
                  + _fmt(cell["value"]))
            for key, cell in summary.items()}


def _collect_amy_rows(rows: list[dict[str, Any]]) -> dict[tuple[str, str], str]:
    # representative per (airline, route_type) = cheapest fare (min net_pay)
    best: dict[tuple[str, str], dict[str, Any]] = {}
    for r in rows:
        rt = _route_type(r["origin"], r["destination"])
        key = (rt, r["airline"])
        if key not in best or r["net_pay"] < best[key]["net_pay"]:
            best[key] = r
    return {key: _fmt(r["commission_pct"]) for key, r in best.items()}


def collect_amy(har_path: str) -> dict[tuple[str, str], str]:
    return _collect_amy_rows(amyweb.parse_agent_har(har_path))


# --- grid assembly + rendering ----------------------------------------------------------

def _oracle_from_rows(ft_b2b_rows: list[dict[str, Any]],
                      b2c_rows: list[dict[str, Any]]):
    """Domestic true-base oracle from ALREADY-PARSED rows, plus a health summary.

    OFFLINE-FIRST: the FT B2B HAR rows alone are enough (no network); the live B2C
    rows only enrich coverage. Returns (TrueBase, health) where health =
    {source: ft_b2b_har|live_b2c|ft_b2b_har+live_b2c|none, airlines_covered,
    sample_count} — callers must treat source=="none" as NOT normalized.
    """
    from modules import true_base as tb_mod
    tb = tb_mod.build_from_rows(ft_b2b_rows=ft_b2b_rows, ft_b2c_rows=b2c_rows)

    def _dom_points(rows: list[dict[str, Any]]) -> int:
        return sum(1 for r in rows
                   if tb_mod.is_domestic(str(r.get("origin", "")), str(r.get("destination", ""))))

    b2b_pts, b2c_pts = _dom_points(ft_b2b_rows), _dom_points(b2c_rows)
    if tb.is_empty():
        source = "none"
    elif b2b_pts and b2c_pts:
        source = "ft_b2b_har+live_b2c"
    elif b2b_pts:
        source = "ft_b2b_har"
    else:
        source = "live_b2c"
    health = {
        "source": source,
        "airlines_covered": sorted(tb.by_gross),
        "sample_count": sum(len(g) for g in tb.by_gross.values()),
    }
    return tb, health


def _build_true_base(firsttrip_b2b_hars: Optional[list[str]],
                     routes: list[tuple[str, str, Optional[str]]], date: Optional[str]):
    """Convenience oracle builder from paths/routes (parses + fetches itself).
    NOTE: build_report() does NOT use this — it parses/fetches once and shares the
    rows with the collectors via _oracle_from_rows() (no double work)."""
    ft_rows: list[dict[str, Any]] = []
    for h in (firsttrip_b2b_hars or []):
        ft_rows += firsttrip.parse_b2b_commissions(h)
    b2c_rows = [r for rows in _fetch_firsttrip_b2c(routes, date).values() for r in rows]
    tb, _health = _oracle_from_rows(ft_rows, b2c_rows)
    return tb


def _merge_cells(dicts: list[dict[tuple[str, str], str]]) -> dict[tuple[str, str], str]:
    """Union cells from several HARs of one channel; first non-empty value per key wins.
    Lets short, eviction-safe captures be combined into full coverage."""
    merged: dict[tuple[str, str], str] = {}
    for d in dicts:
        for key, val in d.items():
            if val and key not in merged:
                merged[key] = val
    return merged


def _apply_manual_overrides(channel_cells: dict[str, dict[tuple[str, str], str]],
                            sources: dict[str, str],
                            overrides: dict[str, Any]) -> None:
    """Set explicit cell values for channels that can't be captured (e.g. ShareTrip
    when the live/HAR search fails). Shape: {row_label: {DOM|INTL: {airline: "value"}}}.
    Manual values are authoritative for the cells they name; other cells are untouched."""
    for label, rt_map in overrides.items():
        if label.startswith("_") or not isinstance(rt_map, dict):
            continue  # skip metadata keys like "_comment"
        cell_map = channel_cells.setdefault(label, {})
        n = 0
        for rt, air_map in (rt_map or {}).items():
            rt = rt.upper()
            for airline, text in (air_map or {}).items():
                if text in (None, ""):
                    continue
                cell_map[(rt, airline.upper())] = str(text)
                n += 1
        if n:
            existing = sources.get(label)
            tag = f"manual: {n} cell(s)"
            sources[label] = f"{existing} | {tag}" if existing else tag


def build_report(date: Optional[str], routes: list[tuple[str, str, Optional[str]]],
                 gozayaan_hars: Optional[list[str]] = None, amy_hars: Optional[list[str]] = None,
                 firsttrip_b2b_hars: Optional[list[str]] = None,
                 sharetrip_hars: Optional[list[str]] = None,
                 akij_hars: Optional[list[str]] = None,
                 bdfare_hars: Optional[list[str]] = None,
                 firsttrip_b2c_hars: Optional[list[str]] = None,
                 manual_overrides: Optional[dict[str, Any]] = None,
                 use_true_base: bool = True,
                 run_dt: Optional[datetime] = None) -> dict[str, Any]:
    channel_cells: dict[str, dict[tuple[str, str], str]] = {}
    sources: dict[str, str] = {}

    # Parse each FT B2B HAR ONCE and live-fetch each FT B2C route ONCE; the rows feed
    # BOTH the true-base oracle and the channel collectors (no double parse/fetch —
    # a full HAR parse and a live fetch each used to run twice per report).
    ft_b2b_rows_per_har = [firsttrip.parse_b2b_commissions(h)
                           for h in (firsttrip_b2b_hars or [])]
    b2c_rows_by_route = _fetch_firsttrip_b2c(routes, date) if routes else {}

    # Amy rows parsed once too: they feed the Amy cells AND the market base index
    # BDFare borrows exact base fares from (same flight+date+time+gross).
    amy_rows_by_path: dict[str, list[dict[str, Any]]] = {}
    for h in (amy_hars or []):
        try:
            amy_rows_by_path[h] = amyweb.parse_agent_har(h)
        except Exception:   # noqa: BLE001 — add() replays per-file errors below
            pass
    base_index = _market_base_index(
        [("Amy", rows) for rows in amy_rows_by_path.values()]
        + [("USBA FT-B2B", rows) for rows in ft_b2b_rows_per_har])

    # FT B2C HAR FAILSAFE: when the live fetch is blocked (offline/CF-challenged)
    # or wasn't configured, a saved b2c-api.firsttrip.com search HAR supplies the
    # same rows. Live wins per route; the HAR only fills routes live didn't cover.
    b2c_har_used: list[str] = []
    if firsttrip_b2c_hars:
        live_covered = {(o, d) for (o, d, _dt), rws in b2c_rows_by_route.items() if rws}
        for h in firsttrip_b2c_hars:
            try:
                har_rows = firsttrip.parse_b2c_har(h)
            except Exception as exc:  # noqa: BLE001
                print(f"  ! FT B2C HAR {Path(h).name}: {exc}")
                continue
            groups: dict[tuple[str, str], list[dict[str, Any]]] = {}
            for r in har_rows:
                groups.setdefault((r["origin"], r["destination"]), []).append(r)
            added = 0
            for (o, d), rws in groups.items():
                if (o, d) not in live_covered:
                    # UNION across multiple failsafe HARs for the same route, don't
                    # overwrite — a second B2C HAR should add coverage, not replace it.
                    b2c_rows_by_route.setdefault((o, d, "har"), []).extend(rws)
                    added += len(rws)
            if added:
                b2c_har_used.append(Path(h).name)
                print(f"  FT B2C failsafe: {added} offers from {Path(h).name}")

    # True-base oracle: BDFare/AKIJ domestic cells get recomputed on the canonical base
    # so the grid shows the ACTUAL % (on by default; --no-true-base opts out). Built
    # OFFLINE-FIRST from the FT B2B HAR rows; live B2C rows only enrich coverage. If
    # no exact-base source exists, fall back to each channel's own base rather than
    # blanking cells — and record it in report["true_base"] / report["normalized"].
    tb_health: dict[str, Any] = {"source": "disabled", "airlines_covered": [], "sample_count": 0}
    true_base = None
    if use_true_base:
        oracle, tb_health = _oracle_from_rows(
            [r for rows in ft_b2b_rows_per_har for r in rows],
            [r for rows in b2c_rows_by_route.values() for r in rows])
        if tb_health["source"] == "none":
            print("  ! true-base: no exact-base source (FT B2B HAR / domestic route) — "
                  "BDFare/AKIJ kept on their own base (NOT normalized)")
        else:
            true_base = oracle
    tb_tag = "  [true-base]" if true_base is not None else ""

    # Per-channel capture status: "ok" | "captured_but_empty" (a HAR was provided but
    # parsed to zero cells — e.g. the ShareTrip body-eviction trap) | "parse_failed"
    # (a HAR raised, e.g. a truncated DevTools export) | "manual" | "not_attempted".
    channel_status: dict[str, str] = {}

    def add(label: str, hars: Optional[list[str]], collector, prefix: str) -> None:
        if not hars:
            return
        # Isolate each HAR: one corrupt/truncated export must not kill the whole run
        # (the FT B2C paths already do this). Partial results from the others still
        # merge, and the failure is surfaced via channel_status + the run log.
        parsed, failed = [], []
        for h in hars:
            try:
                parsed.append(collector(h))
            except Exception as exc:  # noqa: BLE001
                failed.append(Path(h).name)
                print(f"  ! {label}: could not parse {Path(h).name}: {exc}")
        channel_cells[label] = _merge_cells(parsed)
        sources[label] = f"{prefix}: {', '.join(Path(h).name for h in hars)}"
        if channel_cells[label]:
            channel_status[label] = "ok"
        elif failed:
            channel_status[label] = "parse_failed"
        else:
            channel_status[label] = "captured_but_empty"

    add("AKIJ AIR-B2B", akij_hars,
        lambda h: collect_akij(h, "realized_discount_pct", true_base), "HAR")
    add("BDFare", bdfare_hars, lambda h: collect_bdfare(h, true_base, base_index), "HAR")
    if true_base is not None:
        for lbl in ("AKIJ AIR-B2B", "BDFare"):
            if lbl in sources:
                sources[lbl] += tb_tag
    if ft_b2b_rows_per_har:
        channel_cells["USBA OTA B2B"] = _merge_cells(
            [_collect_firsttrip_b2b_rows(rows) for rows in ft_b2b_rows_per_har])
        sources["USBA OTA B2B"] = ("FT-B2B HAR: "
                                   + ", ".join(Path(h).name for h in firsttrip_b2b_hars))
        channel_status["USBA OTA B2B"] = ("ok" if channel_cells["USBA OTA B2B"]
                                          else "captured_but_empty")
    if sharetrip_hars:
        # Whole-list (not per-file) so a booking capture's uniform card special can
        # enrich a separate search capture's airlines — search + one booking covers all.
        channel_cells["ShareTrip-B2C"] = collect_sharetrip_b2c(sharetrip_hars)
        sources["ShareTrip-B2C"] = "HAR: " + ", ".join(Path(h).name for h in sharetrip_hars)
        channel_status["ShareTrip-B2C"] = ("ok" if channel_cells["ShareTrip-B2C"]
                                           else "captured_but_empty")
    add("Go Zayaan", gozayaan_hars, collect_gozayaan, "HAR")
    add("Amy", amy_hars,
        lambda h: _collect_amy_rows(amy_rows_by_path[h])
        if h in amy_rows_by_path else collect_amy(h), "HAR")

    if routes or b2c_rows_by_route:
        channel_cells["Firsttrip-B2C"] = _collect_firsttrip_b2c_rows(b2c_rows_by_route)
        parts = []
        if routes:
            parts.append(f"live: {len(routes)} route(s)")
        if b2c_har_used:
            parts.append(f"HAR failsafe: {', '.join(b2c_har_used)}")
        sources["Firsttrip-B2C"] = " | ".join(parts)
        channel_status["Firsttrip-B2C"] = ("ok" if channel_cells["Firsttrip-B2C"]
                                           else "captured_but_empty")

    if manual_overrides:
        _apply_manual_overrides(channel_cells, sources, manual_overrides)
        for label, rt_map in manual_overrides.items():
            if not label.startswith("_") and isinstance(rt_map, dict) \
                    and channel_status.get(label) in (None, "captured_but_empty"):
                channel_status[label] = "manual"
    for label, _kind in ROW_ORDER:
        if _kind != "sep":
            channel_status.setdefault(label, "not_attempted")

    # which airlines actually appear, per route_type
    present = {"DOM": set(), "INTL": set()}
    for cells in channel_cells.values():
        for (rt, airline) in cells:
            present[rt].add(airline)

    def columns(rt: str, base: list[str]) -> list[str]:
        extras = sorted(a for a in present[rt] if a not in base)
        return [a for a in base if a in present[rt]] + extras

    grids = {}
    for rt, base in (("DOM", DOM_COLUMNS), ("INTL", INTL_COLUMNS)):
        cols = columns(rt, base)
        rows = []
        for label, kind in ROW_ORDER:
            if kind == "sep":
                rows.append({"label": "__sep__", "kind": "sep", "cells": {}})
                continue
            cells = channel_cells.get(label, {})
            rows.append({
                "label": label, "kind": kind,
                "cells": {a: cells.get((rt, a), "") for a in cols},
            })
        grids[rt] = {"columns": cols, "rows": rows}

    now = run_dt or datetime.now()
    return {
        "generated_at": now.isoformat(),
        "report_date": now.strftime("%d/%m/%Y"),   # report run date (header), not a travel date
        "report_time": now.strftime("%H%M"),
        "default_date": date,                       # default travel date for live FT B2C, if any
        "routes": [f"{o}-{d}" + (f"@{dt}" if dt else "") for o, d, dt in routes],
        "sources": sources,
        "true_base": tb_health,                     # oracle health: source/coverage
        "normalized": true_base is not None,        # False => BDFare/AKIJ on their own base
        "channel_status": channel_status,           # ok|captured_but_empty|manual|not_attempted
        "grids": grids,
    }


def render_console(report: dict[str, Any]) -> None:
    print("\n" + "=" * 70)
    print(f"OTA DISCOUNT GRID   {report['report_date']} / {report['report_time']}hrs")
    print("=" * 70)
    for ch, src in report["sources"].items():
        print(f"  {ch:<16} {src}")
    if report.get("true_base", {}).get("source") == "none":
        print("  !! NOT NORMALIZED: no exact-base source — BDFare/AKIJ shown on their own base")

    for rt, title in (("DOM", "DOMESTIC"), ("INTL", "INTERNATIONAL")):
        grid = report["grids"][rt]
        cols = grid["columns"]
        if not cols:
            continue
        w = 16
        print(f"\n{title}")
        header = f"{'OTA':<18}" + "".join(f"{c:>{w}}" for c in cols)
        print(header)
        print("-" * len(header))
        for row in grid["rows"]:
            if row["kind"] == "sep":
                print("." * len(header))
                continue
            line = f"{row['label']:<18}" + "".join(f"{row['cells'][c] or '-':>{w}}" for c in cols)
            print(line)
    print("=" * 70)


def _sidebyside_table(report: dict[str, Any]) -> list[list[str]]:
    """The INTERNATIONAL (left) / DOMESTIC (right) table as a list of row-lists.

    One blank spacer column between the two blocks, sharing the OTA row labels.
    Shared by the CSV and XLSX writers so both stay identical. Separator rows
    are returned as an empty list (`[]`).
    """
    stamp_label = f"{report['report_date']}"
    time_label = f"{report['report_time']}hrs"
    intl = report["grids"]["INTL"]
    dom = report["grids"]["DOM"]
    intl_cols, dom_cols = intl["columns"], dom["columns"]
    intl_rows = {r["label"]: r for r in intl["rows"]}
    dom_rows = {r["label"]: r for r in dom["rows"]}

    table: list[list[str]] = []
    table.append([f"{stamp_label} (INTERNATIONAL)/ {time_label}"] + [""] * len(intl_cols)
                 + [""] + [f"{stamp_label} (DOMESTIC)/ {time_label}"] + [""] * len(dom_cols))
    table.append(["OTA"] + intl_cols + [""] + ["OTA"] + dom_cols)
    for label, kind in ROW_ORDER:
        if kind == "sep":
            table.append([])
            continue
        intl_cells = intl_rows.get(label, {}).get("cells", {})
        dom_cells = dom_rows.get(label, {}).get("cells", {})
        table.append([label] + [intl_cells.get(c, "") for c in intl_cols]
                     + [""] + [label] + [dom_cells.get(c, "") for c in dom_cols])
    return table


def write_outputs(report: dict[str, Any], out_dir: Path, stamp: str) -> list[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    base = f"ota_discount_grid_{stamp}"
    paths = []

    json_path = out_dir / f"{base}.json"
    json_path.write_text(json.dumps(report, indent=2, default=str), encoding="utf-8")
    paths.append(json_path)

    csv_path = out_dir / f"{base}.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for row in _sidebyside_table(report):
            w.writerow(row)
    paths.append(csv_path)
    return paths


def _sheet_name(report: dict[str, Any]) -> str:
    """Daily sheet name in the manual report's 'DD Month' convention (e.g. '29 June')."""
    run = datetime.strptime(report["report_date"], "%d/%m/%Y")
    return run.strftime("%d %B")


# Best-discount highlight palette (per airline column, within each B2B / B2C block).
# Highlight palette — fill + text colour (professional, high-contrast; Excel Good/Neutral/Bad style).
HL_GREEN, HL_GREEN_TX = "C6EFCE", "006100"   # best per airline
HL_BLUE, HL_BLUE_TX = "DDEBF7", "1F4E79"     # second
HL_RED, HL_RED_TX = "FFC7CE", "9C0006"       # changed vs the previous report
HDR_BG, HDR_TX = "1F4E79", "FFFFFF"          # header band (navy)
BEST_BG, BEST_TX = "1F2937", "FFFFFF"        # "Best" summary row (slate)
LEGEND = "Green = Best   ·   Blue = 2nd   ·   Red = Change"
# BEST_SHORT (short OTA labels for the "Best" row) now lives in discount_engine.highlight.


def _find_prev_sheet(wb, today_name: str, year: int):
    """The existing sheet with the latest 'DD Month' date strictly before today's
    (for the Red=Change diff). Returns a worksheet or None."""
    def parse(n: str):
        try:
            return datetime.strptime(f"{n} {year}", "%d %B %Y").date()
        except ValueError:
            return None
    today = parse(today_name)
    if today is None:
        return None
    best_date = best_ws = None
    for name in wb.sheetnames:
        if name == today_name:
            continue
        d = parse(name)
        if d is not None and d < today and (best_date is None or d > best_date):
            best_date, best_ws = d, wb[name]
    return best_ws


def _read_prev_grid(ws) -> dict[tuple[str, str, str], float]:
    """Read a prior colored sheet into {(rt, ota_label, airline): fraction}."""
    grid: dict[tuple[str, str, str], float] = {}
    rt = None
    headers: list[str] = []
    for row in ws.iter_rows():
        a = row[0].value
        if isinstance(a, str) and "(INTERNATIONAL)" in a:
            rt, headers = "INTL", []
            continue
        if isinstance(a, str) and "(DOMESTIC)" in a:
            rt, headers = "DOM", []
            continue
        if isinstance(a, str) and a.strip() == "OTA":
            headers = [str(c.value).strip() if c.value is not None else "" for c in row]
            continue
        if rt and headers and isinstance(a, str) and a.strip():
            for ci in range(1, len(headers)):
                air = headers[ci]
                if not air:
                    continue
                v = row[ci].value if ci < len(row) else None
                if isinstance(v, (int, float)):
                    grid[(rt, a.strip(), air)] = float(v)
                elif isinstance(v, str):
                    n = _leading_number(v)   # coupon cells are text; rank by the common rate
                    if n is not None:
                        grid[(rt, a.strip(), air)] = n / 100.0
    return grid


def _render_report_sheet(ws, report: dict[str, Any],
                         prev_lookup: dict[tuple[str, str, str], float]) -> None:
    """Render the colored grid into a worksheet, coloring from compute_highlights().

    Layout matches the manual Commission.xlsx colored sheets: OTA names as ROWS,
    airlines as COLUMNS, INTERNATIONAL and DOMESTIC as stacked blocks, each with a
    merged title (date/time + legend), a navy header row, per-airline green/blue/red
    highlighting within each B2B / B2C group, and the slate "Best (OTA)" summary row.
    Coupon cells show BOTH the common and best-card rates in the cell itself.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    hl = compute_highlights(report, prev_lookup)

    cal = "Calibri"
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor=HDR_BG)
    best_fill = PatternFill("solid", fgColor=BEST_BG)
    title_font = Font(name=cal, bold=True, size=11)
    head_font = Font(name=cal, bold=True, size=11, color=HDR_TX)
    label_font = Font(name=cal, bold=True, size=11)
    data_font = Font(name=cal, size=11)
    best_font = Font(name=cal, bold=True, size=11, color=BEST_TX)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    flag_style = {   # highlight flag -> (fill, font)
        "changed": (PatternFill("solid", fgColor=HL_RED), Font(name=cal, size=11, color=HL_RED_TX)),
        "highest": (PatternFill("solid", fgColor=HL_GREEN),
                    Font(name=cal, bold=True, size=11, color=HL_GREEN_TX)),
        "second": (PatternFill("solid", fgColor=HL_BLUE), Font(name=cal, size=11, color=HL_BLUE_TX)),
    }

    date_label = report["report_date"]
    time_label = f"{report['report_time']}hrs"
    max_cols = 1

    r = 1
    for rt, name in (("INTL", "INTERNATIONAL"), ("DOM", "DOMESTIC")):
        grid = report.get("grids", {}).get(rt)   # tolerate single-block reports
        if not grid:
            continue
        cols = grid["columns"]
        if not cols:
            continue
        ncol = 1 + len(cols)
        max_cols = max(max_cols, ncol)
        rows_by_label = {row["label"]: row.get("cells", {}) for row in grid["rows"]}
        flags, best = hl[rt]["flags"], hl[rt]["best"]

        # Title row (merged across the block) with the legend.
        tcell = ws.cell(r, 1, f"{date_label} ({name})/ {time_label}\n{LEGEND}")
        tcell.font, tcell.alignment = title_font, center
        for ci in range(1, ncol + 1):
            ws.cell(r, ci).border = border
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
        ws.row_dimensions[r].height = 30
        r += 1

        # Header row (navy band): OTA + airline codes.
        ws.cell(r, 1, "OTA")
        for ci, airline in enumerate(cols, start=2):
            ws.cell(r, ci, airline)
        for ci in range(1, ncol + 1):
            c = ws.cell(r, ci)
            c.font, c.fill, c.alignment, c.border = head_font, hdr_fill, center, border
        r += 1

        # Data rows: values written here; fills/fonts come from the shared flags.
        for label, kind in ROW_ORDER:
            if kind == "sep":
                r += 1
                continue
            cells = rows_by_label.get(label, {})
            lc = ws.cell(r, 1, label)
            lc.font, lc.alignment, lc.border = label_font, left, border
            has_coupon = False
            for ci, airline in enumerate(cols, start=2):
                raw = str(cells.get(airline, "") or "").strip()
                cell = ws.cell(r, ci)
                cell.font, cell.alignment, cell.border = data_font, right, border
                num = _leading_number(raw)
                if num is None:
                    cell.value = raw or None
                elif _fmt(num) == raw:
                    # plain single-rate cell -> number with % format
                    cell.value, cell.number_format = num / 100.0, "0.00%"
                else:
                    # coupon cell (common + special) -> show BOTH %s in the cell
                    cell.value = re.sub(r"(\d+(?:\.\d+)?)", r"\1%", raw)
                    cell.alignment = center   # wrap the longer coupon text
                    has_coupon = True
                style = flag_style.get(flags.get((label, airline)))
                if style:
                    cell.fill, cell.font = style
            if has_coupon:
                ws.row_dimensions[r].height = 30
            r += 1

        # "Best (OTA)" summary row (slate band) from the shared best-per-airline.
        blabel = ws.cell(r, 1, "Best (OTA)")
        blabel.font, blabel.fill, blabel.alignment, blabel.border = best_font, best_fill, left, border
        for ci, airline in enumerate(cols, start=2):
            cell = ws.cell(r, ci)
            cell.font, cell.fill, cell.alignment, cell.border = best_font, best_fill, center, border
            if airline in best:
                cell.value = best[airline]["display"]
        ws.row_dimensions[r].height = 30
        r += 3   # Best row + gap between the INTERNATIONAL and DOMESTIC blocks

    ws.column_dimensions["A"].width = 20
    for ci in range(2, max_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 12
    ws.freeze_panes = "B1"   # keep OTA labels visible when scrolling across airlines


def _save_workbook(wb, xlsx_path: Path) -> Path:
    """Save; on a Windows Excel lock, write a timestamped sibling instead of dying."""
    try:
        wb.save(xlsx_path)
        return xlsx_path
    except PermissionError:
        fallback = xlsx_path.with_name(
            f"{xlsx_path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{xlsx_path.suffix}")
        wb.save(fallback)
        print(f"  ! {xlsx_path.name} is open/locked — wrote {fallback.name} instead. "
              f"Close the workbook to let the daily sheet update in place.")
        return fallback


def write_xlsx(report: dict[str, Any], xlsx_path: Path) -> Path:
    """Append the colored daily sheet to one persistent (rolling) workbook.

    The red change-diff compares against the latest earlier 'DD Month' sheet in the
    SAME workbook (CLI/desktop rolling-file mode). Re-running the same day overwrites
    that day's sheet. For a standalone file diffed against a stored previous report,
    use write_single_sheet_xlsx().
    """
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise SystemExit(
            "openpyxl is required for the daily Excel sheet "
            "(`pip install openpyxl`), or pass --no-xlsx to skip."
        ) from exc

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    sheet_name = _sheet_name(report)
    year = datetime.strptime(report["report_date"], "%d/%m/%Y").year
    prev_ws = _find_prev_sheet(wb, sheet_name, year)
    # _read_prev_grid returns fractions; the highlight flags work in percent units.
    prev_lookup = {k: v * 100 for k, v in _read_prev_grid(prev_ws).items()} \
        if prev_ws is not None else {}

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]          # overwrite today's sheet on re-run
    ws = wb.create_sheet(title=sheet_name)
    _render_report_sheet(ws, report, prev_lookup)
    return _save_workbook(wb, xlsx_path)


def write_single_sheet_xlsx(report: dict[str, Any],
                            prev_report: Optional[dict[str, Any]],
                            xlsx_path: Path) -> Path:
    """One standalone workbook with just this report's sheet; the red change-diff
    comes from `prev_report` (the STORED previous report dict, e.g. from the backend)
    — never from other sheets. This is the server-regen / web-download writer."""
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise SystemExit("openpyxl is required for xlsx output.") from exc

    xlsx_path = Path(xlsx_path)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title=_sheet_name(report))
    _render_report_sheet(ws, report, prev_lookup_from_report(prev_report))
    return _save_workbook(wb, xlsx_path)


# Substrings that identify which channel a HAR belongs to (first match wins).
HAR_SIGNATURES: list[tuple[str, str]] = [
    ("firsttrip_b2c", "b2c-api.firsttrip.com"),   # before b2b: hostname is specific
    ("firsttrip_b2b", "/api/Search/Progressive"),
    ("bdfare", "/bdfare-search/api/"),
    ("amy", "amyx.amybd.com"),
    ("akij", "akijair.com"),
    ("gozayaan", "/api/business_rules/get_discount_list/"),
    ("gozayaan", "production.gozayaan.com/api/flight"),
    ("sharetrip", "/flight/search/available-flights"),
]

# Cap the content sniff: request URLs appear early in a HAR, and reading a 500 MB
# capture whole just to identify it was the original memory bug (review P0).
_SNIFF_BYTES = 4_000_000


# Filename hints take precedence (collector names files by channel; downloads are named by site).
FILENAME_HINTS: list[tuple[str, str]] = [
    ("akij", "akij"),
    ("bdfare", "bdfare"),
    ("gozayaan", "gozayaan"),
    ("gozayaan", "gozyaan"),
    ("gozayaan", "goz"),  # catch typos (gozyaaaan, gozaayan, …); no other channel filename has "goz"
    ("sharetrip", "sharetrip"),
    ("firsttrip_b2c", "b2c-api.firsttrip"),
    ("firsttrip_b2c", "firsttrip_b2c"),
    ("firsttrip_b2b", "booking.firsttrip"),
    ("firsttrip_b2b", "firsttrip_b2b"),
    ("amy", "amyweb"),
    ("amy", "amybd"),
]


def detect_channel(har_path: Path) -> Optional[str]:
    """Identify a HAR's channel by filename first, then by sniffing the head."""
    name = har_path.name.lower()
    for channel, hint in FILENAME_HINTS:
        if hint in name:
            return channel
    try:
        with open(har_path, encoding="utf-8", errors="ignore") as f:
            text = f.read(_SNIFF_BYTES)
    except OSError:
        return None
    for channel, needle in HAR_SIGNATURES:
        if needle in text:
            return channel
    return None


def auto_detect_hars(har_dir: Path) -> dict[str, list[str]]:
    """Map channel -> [har paths] for every .har in a folder (multiple per channel merge)."""
    found: dict[str, list[str]] = {}
    for har in sorted(har_dir.glob("*.har")):
        channel = detect_channel(har)
        if channel:
            found.setdefault(channel, []).append(str(har))
            print(f"  detected {channel:<14} <- {har.name}")
        else:
            print(f"  (unrecognized) {har.name}")
    return found


def _parse_routes(value: str) -> list[tuple[str, str, Optional[str]]]:
    """Parse 'ORIGIN-DEST' or 'ORIGIN-DEST@YYYY-MM-DD' (per-route date) entries."""
    routes = []
    for item in value.split(","):
        item = item.strip()
        if not item:
            continue
        date = None
        if "@" in item:
            item, date = item.split("@", 1)
            date = date.strip() or None
        item = item.strip().upper()
        if "-" not in item:
            raise SystemExit(f"--routes entries must be ORIGIN-DEST[@DATE], got {item!r}")
        o, d = item.split("-", 1)
        routes.append((o, d, date))
    return routes
