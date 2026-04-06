from __future__ import annotations

import io
import json
from datetime import UTC, date, datetime
from typing import Any, Sequence

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from . import reporting


EXPORT_SECTION_ORDER = ("routes", "operations", "changes", "taxes", "penalties")
AIRLINE_THEME: dict[str, dict[str, str]] = {
    "BG": {"header": "C8102E", "sub": "FFF1F4", "text": "5F1020", "header_text": "FFFFFF"},
    "VQ": {"header": "003A70", "sub": "FFE7D0", "text": "123A6E", "header_text": "FFFFFF"},
    "BS": {"header": "00557F", "sub": "D8EBF7", "text": "11384D", "header_text": "FFFFFF"},
    "2A": {"header": "B78700", "sub": "FDEFC7", "text": "6B4E00", "header_text": "1E1E1E"},
}
DEFAULT_THEME = {"header": "194866", "sub": "DCECF6", "text": "163449", "header_text": "FFFFFF"}
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
ROUTE_FILL = PatternFill("solid", fgColor="E6F0F7")
META_FILL = PatternFill("solid", fgColor="F7F4EE")


def _json_ready(value: Any) -> Any:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=True)
    return value


def _rows_to_frame(rows: list[dict[str, Any]]) -> pd.DataFrame:
    if not rows:
        return pd.DataFrame()
    normalized = [{key: _json_ready(value) for key, value in row.items()} for row in rows]
    return pd.DataFrame(normalized)


def _flatten_route_monitor(payload: dict[str, Any]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for route in payload.get("routes", []):
        flight_lookup = {
            str(item.get("flight_group_id")): item
            for item in route.get("flight_groups", [])
        }
        for date_group in route.get("date_groups", []):
            captures = list(date_group.get("captures") or [])
            for capture_index, capture in enumerate(captures):
                for cell in capture.get("cells", []):
                    flight = flight_lookup.get(str(cell.get("flight_group_id"))) or {}
                    rows.append(
                        {
                            "cycle_id": payload.get("cycle_id"),
                            "route_key": route.get("route_key"),
                            "origin": route.get("origin"),
                            "destination": route.get("destination"),
                            "route_type": route.get("route_type"),
                            "origin_country_code": route.get("origin_country_code"),
                            "destination_country_code": route.get("destination_country_code"),
                            "country_pair": route.get("country_pair"),
                            "domestic_country_code": route.get("domestic_country_code"),
                            "is_cross_border": route.get("is_cross_border"),
                            "search_trip_type": route.get("search_trip_type"),
                            "trip_pair_key": route.get("trip_pair_key"),
                            "requested_outbound_date": route.get("requested_outbound_date"),
                            "requested_return_date": route.get("requested_return_date"),
                            "trip_duration_days": route.get("trip_duration_days"),
                            "trip_origin": route.get("trip_origin"),
                            "trip_destination": route.get("trip_destination"),
                            "departure_date": date_group.get("departure_date"),
                            "day_label": date_group.get("day_label"),
                            "captured_at_utc": capture.get("captured_at_utc"),
                            "is_latest_capture": capture_index == 0,
                            "airline": flight.get("airline"),
                            "flight_number": flight.get("flight_number"),
                            "departure_time": flight.get("departure_time"),
                            "cabin": flight.get("cabin"),
                            "aircraft": flight.get("aircraft"),
                            "leg_direction": flight.get("leg_direction"),
                            "leg_sequence": flight.get("leg_sequence"),
                            "itinerary_leg_count": flight.get("itinerary_leg_count"),
                            "signal": cell.get("signal"),
                            "min_total_price_bdt": cell.get("min_total_price_bdt"),
                            "max_total_price_bdt": cell.get("max_total_price_bdt"),
                            "tax_amount": cell.get("tax_amount"),
                            "booking_class": cell.get("booking_class"),
                            "seat_available": cell.get("seat_available"),
                            "seat_capacity": cell.get("seat_capacity"),
                            "load_factor_pct": cell.get("load_factor_pct"),
                            "soldout": cell.get("soldout"),
                        }
                    )
    if not rows:
        return pd.DataFrame()
    frame = pd.DataFrame(rows)
    return frame.sort_values(
        by=["route_key", "departure_date", "captured_at_utc", "departure_time", "airline", "flight_number"],
        ascending=[True, True, False, True, True, True],
        na_position="last",
    )


def _theme_for_airline(code: str) -> dict[str, str]:
    return AIRLINE_THEME.get(str(code or "").strip().upper(), DEFAULT_THEME)


def _format_money_value(value: Any) -> str:
    if value is None or value == "":
        return "—"
    try:
        return f"{float(value):,.0f}"
    except Exception:
        return str(value)


def _format_percent_value(value: Any) -> str:
    if value is None or value == "":
        return "—"
    try:
        return f"{float(value):.1f}%"
    except Exception:
        return str(value)


def _format_capture_label(value: Any) -> str:
    if not value:
        return "-"
    try:
        dt = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        return dt.strftime("%d %b %Y, %H:%M")
    except Exception:
        return str(value)


def _fare_signal_tag(cell: dict[str, Any]) -> str | None:
    if cell.get("soldout") or cell.get("signal") == "sold_out":
        return "SOLD OUT"
    if cell.get("signal") == "new":
        return "NEW"
    return None


def _fare_display(cell: dict[str, Any] | None) -> str:
    if not cell:
        return "N/O"

    min_fare = "N/O" if cell.get("min_total_price_bdt") is None else _format_money_value(cell.get("min_total_price_bdt"))
    signal = str(cell.get("signal") or "").strip().lower()
    if signal == "increase":
        min_fare = f"{min_fare} ↑"
    elif signal == "decrease":
        min_fare = f"{min_fare} ↓"

    details: list[str] = []
    booking_class = str(cell.get("booking_class") or "").strip()
    if booking_class:
        details.append(booking_class)
    if cell.get("seat_available") is not None:
        seats = int(cell.get("seat_available"))
        details.append(f"{seats} seat" if seats == 1 else f"{seats} seats")
    tag = _fare_signal_tag(cell)
    parts = [min_fare]
    if details:
        parts.append(" · ".join(details))
    if tag:
        parts.append(tag)
    return "\n".join(parts)


def _route_leader(route: dict[str, Any]) -> str:
    flight_lookup = {
        str(item.get("flight_group_id")): item
        for item in route.get("flight_groups", [])
    }
    best: dict[str, Any] | None = None
    best_dates: list[str] = []
    for date_group in route.get("date_groups", []):
        captures = date_group.get("captures") or []
        latest_capture = captures[0] if captures else None
        if not latest_capture:
            continue
        for cell in latest_capture.get("cells", []):
            amount = cell.get("min_total_price_bdt")
            if amount is None:
                continue
            flight = flight_lookup.get(str(cell.get("flight_group_id"))) or {}
            if best is None or float(amount) < float(best["amount"]):
                best = {
                    "amount": amount,
                    "airline": flight.get("airline"),
                    "flight_number": flight.get("flight_number"),
                }
                best_dates = [str(date_group.get("departure_date") or "")]
            elif float(amount) == float(best["amount"]):
                dep_date = str(date_group.get("departure_date") or "")
                if dep_date and dep_date not in best_dates:
                    best_dates.append(dep_date)
    if not best:
        return "No visible fare leader"
    date_label = ", ".join(sorted(best_dates))
    return f"Route Price Leader (Lowest Fare): {best['airline']}{best['flight_number']} {_format_money_value(best['amount'])} (Dates: {date_label})"


def _auto_fit_route_columns(sheet) -> None:
    widths = {
        1: 14,
        2: 12,
        3: 22,
    }
    for idx in range(1, sheet.max_column + 1):
        width = widths.get(idx, 15)
        sheet.column_dimensions[get_column_letter(idx)].width = width


def _write_route_monitor_sheet(writer: pd.ExcelWriter, payload: dict[str, Any]) -> int:
    workbook = writer.book
    sheet = workbook.create_sheet(title="Routes", index=0)
    row_cursor = 1

    for route in payload.get("routes", []):
        flight_groups = list(route.get("flight_groups") or [])
        if not flight_groups:
            continue

        show_seat_column = any(
            cell.get("seat_available") is not None or cell.get("seat_capacity") is not None
            for date_group in route.get("date_groups", [])
            for capture in date_group.get("captures", [])
            for cell in capture.get("cells", [])
        )
        show_load_column = any(
            cell.get("load_factor_pct") is not None
            for date_group in route.get("date_groups", [])
            for capture in date_group.get("captures", [])
            for cell in capture.get("cells", [])
        )

        metric_headers = ["Min Fare", "Max Fare", "Tax Amount"]
        if show_seat_column:
            metric_headers.append("Open/Cap")
        if show_load_column:
            metric_headers.append("Inv Press")
        metric_count = len(metric_headers)
        total_columns = 3 + len(flight_groups) * metric_count

        sheet.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=3)
        route_cell = sheet.cell(row=row_cursor, column=1, value=str(route.get("route_key") or "Route"))
        route_cell.font = Font(size=16, bold=True, color="10233B")
        route_cell.fill = ROUTE_FILL
        route_cell.alignment = Alignment(vertical="center")
        route_cell.border = THIN_BORDER

        sheet.merge_cells(start_row=row_cursor, start_column=4, end_row=row_cursor, end_column=total_columns)
        leader_cell = sheet.cell(row=row_cursor, column=4, value=_route_leader(route))
        leader_cell.font = Font(bold=True, color="163449")
        leader_cell.fill = ROUTE_FILL
        leader_cell.alignment = Alignment(vertical="center")
        leader_cell.border = THIN_BORDER
        row_cursor += 1

        static_headers = ["Date", "Day", "Capture Date/Time"]
        for col_idx, label in enumerate(static_headers, start=1):
            cell = sheet.cell(row=row_cursor, column=col_idx, value=label)
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        column_cursor = 4
        for flight in flight_groups:
            theme = _theme_for_airline(str(flight.get("airline") or ""))
            sheet.merge_cells(
                start_row=row_cursor,
                start_column=column_cursor,
                end_row=row_cursor,
                end_column=column_cursor + metric_count - 1,
            )
            cell = sheet.cell(
                row=row_cursor,
                column=column_cursor,
                value=f"{flight.get('airline')}{flight.get('flight_number')} | {flight.get('aircraft') or 'Flight'}",
            )
            cell.font = Font(bold=True, color=theme["header_text"])
            cell.fill = PatternFill("solid", fgColor=theme["header"])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
            column_cursor += metric_count
        row_cursor += 1

        for col_idx in range(1, 4):
            cell = sheet.cell(row=row_cursor, column=col_idx, value="")
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER

        column_cursor = 4
        for flight in flight_groups:
            theme = _theme_for_airline(str(flight.get("airline") or ""))
            sheet.merge_cells(
                start_row=row_cursor,
                start_column=column_cursor,
                end_row=row_cursor,
                end_column=column_cursor + metric_count - 1,
            )
            cell = sheet.cell(
                row=row_cursor,
                column=column_cursor,
                value=str(flight.get("departure_time") or "—"),
            )
            cell.font = Font(bold=True, color=theme["header_text"])
            cell.fill = PatternFill("solid", fgColor=theme["header"])
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
            column_cursor += metric_count
        row_cursor += 1

        for col_idx in range(1, 4):
            cell = sheet.cell(row=row_cursor, column=col_idx, value="" if col_idx > 1 else "")
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER

        column_cursor = 4
        for flight in flight_groups:
            theme = _theme_for_airline(str(flight.get("airline") or ""))
            for metric in metric_headers:
                cell = sheet.cell(row=row_cursor, column=column_cursor, value=metric)
                cell.font = Font(bold=True, color=theme["text"])
                cell.fill = PatternFill("solid", fgColor=theme["sub"])
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = THIN_BORDER
                column_cursor += 1
        row_cursor += 1

        for date_group in route.get("date_groups", []):
            captures = list(date_group.get("captures") or [])
            for capture_index, capture in enumerate(captures):
                show_date_meta = capture_index == 0
                cells_by_flight = {
                    str(cell.get("flight_group_id")): cell
                    for cell in capture.get("cells", [])
                }

                values = [
                    str(date_group.get("departure_date") or "") if show_date_meta else "",
                    str(date_group.get("day_label") or "") if show_date_meta else "",
                    _format_capture_label(capture.get("captured_at_utc")),
                ]
                for column_index, value in enumerate(values, start=1):
                    cell = sheet.cell(row=row_cursor, column=column_index, value=value)
                    cell.border = THIN_BORDER
                    cell.alignment = Alignment(vertical="top")
                    if column_index == 3:
                        cell.fill = META_FILL

                column_cursor = 4
                for flight in flight_groups:
                    theme = _theme_for_airline(str(flight.get("airline") or ""))
                    cell_payload = cells_by_flight.get(str(flight.get("flight_group_id"))) or {}

                    display_values = [
                        _fare_display(cell_payload),
                        _format_money_value(cell_payload.get("max_total_price_bdt")),
                        _format_money_value(cell_payload.get("tax_amount")),
                    ]
                    if show_seat_column:
                        seat_available = cell_payload.get("seat_available")
                        seat_capacity = cell_payload.get("seat_capacity")
                        display_values.append(
                            f"{seat_available if seat_available is not None else '—'} / {seat_capacity if seat_capacity is not None else '—'}"
                        )
                    if show_load_column:
                        display_values.append(_format_percent_value(cell_payload.get("load_factor_pct")))

                    for value in display_values:
                        cell = sheet.cell(row=row_cursor, column=column_cursor, value=value)
                        cell.fill = PatternFill("solid", fgColor=theme["sub"])
                        cell.font = Font(color=theme["text"], bold=column_cursor == 4 or value == display_values[0])
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                        cell.border = THIN_BORDER
                        column_cursor += 1
                row_cursor += 1

        row_cursor += 1

    _auto_fit_route_columns(sheet)
    sheet.sheet_view.showGridLines = True
    return max(row_cursor - 1, 0)


def _metadata_sheet_rows(
    *,
    sections: Sequence[str],
    cycle_id: str | None,
    airlines: Sequence[str] | None,
    origins: Sequence[str] | None,
    destinations: Sequence[str] | None,
    route_types: Sequence[str] | None,
    trip_types: Sequence[str] | None,
    return_date: date | None,
    return_date_start: date | None,
    return_date_end: date | None,
    cabins: Sequence[str] | None,
    start_date: date | None,
    end_date: date | None,
    domains: Sequence[str] | None,
    change_types: Sequence[str] | None,
    directions: Sequence[str] | None,
    route_limit: int,
    history_limit: int,
    limit: int,
    section_row_counts: dict[str, int],
) -> list[dict[str, Any]]:
    filters = {
        "sections": ", ".join(sections),
        "cycle_id": cycle_id or "",
        "airlines": ", ".join(airlines or ()),
        "origins": ", ".join(origins or ()),
        "destinations": ", ".join(destinations or ()),
        "route_types": ", ".join(route_types or ()),
        "trip_types": ", ".join(trip_types or ()),
        "return_date": return_date.isoformat() if return_date else "",
        "return_date_start": return_date_start.isoformat() if return_date_start else "",
        "return_date_end": return_date_end.isoformat() if return_date_end else "",
        "cabins": ", ".join(cabins or ()),
        "start_date": start_date.isoformat() if start_date else "",
        "end_date": end_date.isoformat() if end_date else "",
        "domains": ", ".join(domains or ()),
        "change_types": ", ".join(change_types or ()),
        "directions": ", ".join(directions or ()),
        "route_limit": route_limit,
        "history_limit": history_limit,
        "limit": limit,
    }
    rows = [
        {"group": "export", "key": "generated_at_utc", "value": datetime.now(UTC).isoformat()},
        {"group": "export", "key": "workbook_type", "value": "operational_filter_export"},
    ]
    rows.extend(
        {"group": "filters", "key": key, "value": value}
        for key, value in filters.items()
    )
    rows.extend(
        {"group": "counts", "key": key, "value": value}
        for key, value in section_row_counts.items()
    )
    return rows


def build_reporting_workbook(
    session: Session | None,
    *,
    sections: Sequence[str],
    cycle_id: str | None = None,
    airlines: Sequence[str] | None = None,
    origins: Sequence[str] | None = None,
    destinations: Sequence[str] | None = None,
    route_types: Sequence[str] | None = None,
    trip_types: Sequence[str] | None = None,
    return_date: date | None = None,
    return_date_start: date | None = None,
    return_date_end: date | None = None,
    cabins: Sequence[str] | None = None,
    start_date: date | None = None,
    end_date: date | None = None,
    domains: Sequence[str] | None = None,
    change_types: Sequence[str] | None = None,
    directions: Sequence[str] | None = None,
    route_limit: int = 8,
    history_limit: int = 12,
    limit: int = 250,
) -> tuple[bytes, str]:
    normalized_sections = [item for item in EXPORT_SECTION_ORDER if item in set(sections)]
    if not normalized_sections:
        normalized_sections = list(EXPORT_SECTION_ORDER)

    workbook = io.BytesIO()
    section_row_counts: dict[str, int] = {}

    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        if "routes" in normalized_sections:
            route_payload = reporting.get_route_monitor_matrix(
                session,
                cycle_id=cycle_id,
                airlines=airlines,
                origins=origins,
                destinations=destinations,
                cabins=cabins,
                trip_types=trip_types,
                return_date=return_date,
                return_date_start=return_date_start,
                return_date_end=return_date_end,
                start_date=start_date,
                end_date=end_date,
                route_limit=route_limit,
                history_limit=history_limit,
            )
            section_row_counts["routes"] = _write_route_monitor_sheet(writer, route_payload)

        if "operations" in normalized_sections:
            operations_payload = reporting.get_airline_operations(
                session,
                cycle_id=cycle_id,
                airlines=airlines,
                origins=origins,
                destinations=destinations,
                route_types=route_types,
                start_date=start_date,
                end_date=end_date,
                route_limit=route_limit,
                trend_limit=history_limit,
            )
            operations_rows: list[dict[str, Any]] = []
            for route in operations_payload.get("routes", []):
                for airline_entry in route.get("airlines", []):
                    operations_rows.append(
                        {
                            "cycle_id": operations_payload.get("cycle_id"),
                            "route_key": route.get("route_key"),
                            "origin": route.get("origin"),
                            "destination": route.get("destination"),
                            "route_type": route.get("route_type"),
                            "origin_country_code": route.get("origin_country_code"),
                            "destination_country_code": route.get("destination_country_code"),
                            "country_pair": route.get("country_pair"),
                            "airline": airline_entry.get("airline"),
                            "flight_instance_count": airline_entry.get("flight_instance_count"),
                            "active_date_count": airline_entry.get("active_date_count"),
                            "first_departure_time": airline_entry.get("first_departure_time"),
                            "last_departure_time": airline_entry.get("last_departure_time"),
                            "departure_times": airline_entry.get("departure_times"),
                            "flight_numbers": airline_entry.get("flight_numbers"),
                            "weekday_profile": airline_entry.get("weekday_profile"),
                            "timeline": airline_entry.get("timeline"),
                        }
                    )
            operations_frame = _rows_to_frame(operations_rows)
            section_row_counts["operations"] = int(len(operations_frame))
            operations_frame.to_excel(writer, index=False, sheet_name="Operations")

        if "changes" in normalized_sections:
            change_rows = reporting.get_change_events(
                session,
                start_date=start_date,
                end_date=end_date,
                airlines=airlines,
                origins=origins,
                destinations=destinations,
                domains=domains,
                change_types=change_types,
                directions=directions,
                limit=limit,
            )
            change_frame = _rows_to_frame(change_rows)
            section_row_counts["changes"] = int(len(change_frame))
            change_frame.to_excel(writer, index=False, sheet_name="Changes")

        if "taxes" in normalized_sections:
            tax_payload = reporting.get_taxes(
                session,
                cycle_id=cycle_id,
                airlines=airlines,
                origins=origins,
                destinations=destinations,
                limit=limit,
            )
            tax_frame = _rows_to_frame(tax_payload.get("rows", []))
            section_row_counts["taxes"] = int(len(tax_frame))
            tax_frame.to_excel(writer, index=False, sheet_name="Taxes")

        if "penalties" in normalized_sections:
            penalty_payload = reporting.get_penalties(
                session,
                cycle_id=cycle_id,
                airlines=airlines,
                origins=origins,
                destinations=destinations,
                limit=limit,
            )
            penalty_frame = _rows_to_frame(penalty_payload.get("rows", []))
            section_row_counts["penalties"] = int(len(penalty_frame))
            penalty_frame.to_excel(writer, index=False, sheet_name="Penalties")

        if "Sheet" in writer.book.sheetnames and len(writer.book.sheetnames) > 1:
            writer.book.remove(writer.book["Sheet"])

    workbook.seek(0)
    stamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
    section_token = "_".join(normalized_sections)
    filename = f"aero_pulse_export_{section_token}_{stamp}.xlsx"
    return workbook.getvalue(), filename
