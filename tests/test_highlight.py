"""Tests for discount_engine.highlight — the single source of truth for grid coloring.

Semantics: rank per airline WITHIN each B2B/B2C group by the NET common rate (pct minus
convenience fee; text coupon cells included); changed-vs-previous (gross) wins over
highest/second; the "Best (net)" row is the best NET universal rate plus a gated card tier.
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from discount_engine.highlight import (  # noqa: E402
    apply_highlights,
    compute_highlights,
    leading_number,
    prev_lookup_from_report,
)


def _report(cells_by_label: dict[str, dict[str, str]], cols=("BS", "2A")) -> dict:
    kinds = {"USBA OTA B2B": "b2b", "BDFare": "b2b",
             "Firsttrip-B2C": "b2c", "ShareTrip-B2C": "b2c"}
    rows = [{"label": lab, "kind": kinds[lab],
             "cells": {c: cells_by_label.get(lab, {}).get(c, "") for c in cols}}
            for lab in kinds]
    return {"report_date": "02/07/2026", "report_time": "1200",
            "grids": {"DOM": {"columns": list(cols), "rows": rows}}}


BASE = {
    "USBA OTA B2B": {"BS": "12", "2A": "13.2"},
    "BDFare": {"BS": "8.02", "2A": "13"},
    "Firsttrip-B2C": {"BS": "16", "2A": "16"},
    "ShareTrip-B2C": {"BS": "9(Bkash), 18 (Stellar)", "2A": "8.5(Bkash), 18 (Stellar)"},
}


def test_leading_number_variants():
    assert leading_number("9(Bkash), 18 (EBL)") == 9.0
    assert leading_number("12") == 12.0
    assert leading_number("-6.49") == -6.49
    assert leading_number("") is None and leading_number(None) is None


def test_highest_and_second_within_each_group():
    hl = compute_highlights(_report(BASE))["DOM"]
    f = hl["flags"]
    # B2B group: USBA 12 > BDFare 8.02 for BS
    assert f[("USBA OTA B2B", "BS")] == "highest"
    assert f[("BDFare", "BS")] == "second"
    # B2C group ranks by the COMMON rate of the coupon TEXT cell (16 > 9)
    assert f[("Firsttrip-B2C", "BS")] == "highest"
    assert f[("ShareTrip-B2C", "BS")] == "second"


def test_changed_wins_over_highest():
    prev = _report({**BASE, "USBA OTA B2B": {"BS": "11", "2A": "13.2"}})
    hl = compute_highlights(_report(BASE), prev_lookup_from_report(prev))["DOM"]
    assert hl["flags"][("USBA OTA B2B", "BS")] == "changed"     # 11 -> 12, beats green
    assert hl["flags"][("USBA OTA B2B", "2A")] == "highest"     # unchanged keeps rank


def test_coupon_text_change_detected_by_common_rate():
    prev = _report({**BASE, "ShareTrip-B2C": {"BS": "8(Bkash), 18 (Stellar)",
                                              "2A": "8.5(Bkash), 18 (Stellar)"}})
    hl = compute_highlights(_report(BASE), prev_lookup_from_report(prev))["DOM"]
    assert hl["flags"][("ShareTrip-B2C", "BS")] == "changed"    # common 8 -> 9
    assert hl["flags"][("ShareTrip-B2C", "2A")] == "second"     # unchanged


def test_best_row_across_all_channels():
    best = compute_highlights(_report(BASE))["DOM"]["best"]
    # primary (universal) = best NET rate anyone gets (no fees in BASE -> net == gross)
    assert best["BS"]["display"] == "16% net · FT-B2C"
    assert best["BS"]["channel"] == "Firsttrip-B2C"
    assert best["BS"]["universal"]["net"] == 16.0
    # gated = best card-special (ShareTrip's "18 (Stellar)")
    assert best["BS"]["gated"]["display"] == "18% net · ST-B2C, Stellar"
    assert best["2A"]["universal"]["gross"] == 16.0


def test_best_ranks_by_net_of_fee():
    # FT common 9% but 2% fee -> 7 net; ShareTrip common 8% but 0.5% fee -> 7.5 net.
    # Net ranking must pick ShareTrip as universal-best despite its LOWER gross rate.
    cells = {"USBA OTA B2B": {"BS": "5", "2A": "5"}, "BDFare": {"BS": "5", "2A": "5"},
             "Firsttrip-B2C": {"BS": "9(Bkash, 2% fee)", "2A": "9(Bkash, 2% fee)"},
             "ShareTrip-B2C": {"BS": "8(bKash, 0.5% fee)", "2A": "8(bKash, 0.5% fee)"}}
    hl = compute_highlights(_report(cells))["DOM"]
    assert hl["best"]["BS"]["universal"]["short"] == "ST-B2C"       # 7.5 net beats 7 net
    assert hl["best"]["BS"]["universal"]["net"] == 7.5
    assert hl["flags"][("ShareTrip-B2C", "BS")] == "highest"        # green by NET
    assert hl["flags"][("Firsttrip-B2C", "BS")] == "second"


def test_fee_only_change_flags_changed_by_net():
    """A change that only moves the convenience fee (gross flat, NET moved) must flag
    'changed' — change detection ranks by NET, consistent with highlight/Best."""
    st_prev = {"BS": "9(Bkash, 1% fee)", "2A": "8.5(Bkash), 18 (Stellar)"}
    st_cur = {"BS": "9(Bkash, 3% fee)", "2A": "8.5(Bkash), 18 (Stellar)"}
    prev = _report({**BASE, "ShareTrip-B2C": st_prev})
    cur = _report({**BASE, "ShareTrip-B2C": st_cur})
    hl = compute_highlights(cur, prev_lookup_from_report(prev))["DOM"]
    assert hl["flags"][("ShareTrip-B2C", "BS")] == "changed"    # net 8 -> 6 though gross stays 9
    assert hl["flags"][("ShareTrip-B2C", "2A")] != "changed"    # identical both days


def test_apply_highlights_embeds_flags_best_and_prev_date():
    prev = _report({**BASE, "USBA OTA B2B": {"BS": "11", "2A": "13.2"}})
    out = apply_highlights(_report(BASE), prev)
    dom = out["grids"]["DOM"]
    row = {r["label"]: r for r in dom["rows"]}["USBA OTA B2B"]
    assert row["highlights"]["BS"] == "changed"
    assert dom["best"]["BS"]["display"] == "16% net · FT-B2C"
    assert out["prev_report_date"] == "02/07/2026"
    # original report untouched (copy semantics)
    src = _report(BASE)
    assert "highlights" not in src["grids"]["DOM"]["rows"][0]


def test_write_single_sheet_xlsx_colors_from_prev_report():
    import tempfile
    import openpyxl
    from discount_engine import write_single_sheet_xlsx

    tmp_dir = Path(tempfile.mkdtemp(prefix="hl_test_"))
    prev = _report({**BASE, "USBA OTA B2B": {"BS": "11", "2A": "13.2"}})
    out = write_single_sheet_xlsx(_report(BASE), prev, tmp_dir / "single.xlsx")
    ws = openpyxl.load_workbook(out)["02 July"]

    def fill(label, col_idx):
        for ri in range(1, ws.max_row + 1):
            if ws.cell(ri, 1).value == label:
                c = ws.cell(ri, col_idx)
                return str(c.fill.fgColor.rgb)[-6:] if c.fill.patternType else None
        return None

    assert fill("USBA OTA B2B", 2) == "FFC7CE"       # changed -> red
    assert fill("Firsttrip-B2C", 2) == "C6EFCE"      # highest -> green
    assert any(ws.cell(ri, 1).value == "Best (net)" for ri in range(1, ws.max_row + 1))


if __name__ == "__main__":
    import subprocess
    raise SystemExit(subprocess.call([sys.executable, "-m", "pytest", __file__, "-q"]))
